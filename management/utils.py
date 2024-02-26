from .models import Apartment, Fees, WaterReadouts
from openpyxl import load_workbook
from io import BytesIO
from django.http import HttpResponse
from django.template.loader import get_template
from xhtml2pdf import pisa
from xhtml2pdf.files import pisaFileObject


def html_to_pdf(template_src, context_dict=None):
    if context_dict is None:
        context_dict = {}
    template = get_template(template_src)
    html = template.render(context_dict)
    result = BytesIO()
    pisaFileObject.getNamedFile = lambda self: self.uri
    pdf = pisa.pisaDocument(BytesIO(html.encode("utf8")), result)
    if not pdf.err:
        return HttpResponse(result.getvalue(), content_type='application/pdf')
    return None


def calculate_water_usage(water_usage):
    """Takes WaterReadouts queryset and returns usage in m3
    :param water_usage: WaterReadouts queryset
    :returns: cold, hot, cold_curr, hot_curr"""
    cold_curr, hot_curr = water_usage[0].get_water_readout
    if water_usage[0].new_cold_water_meter:
        cold_old = water_usage[1].get_cold_water_readout - water_usage[2].get_cold_water_readout
        cold = cold_curr + cold_old
    else:
        cold = cold_curr - water_usage[1].get_cold_water_readout
    if water_usage[0].new_hot_water_meter:
        hot_old = water_usage[1].get_hot_water_readout - water_usage[2].get_hot_water_readout
        hot = hot_curr + hot_old
    else:
        hot = hot_curr - water_usage[1].get_hot_water_readout

    return cold, hot, cold_curr, hot_curr


def calculate_fees(apt, water_id=None):

    if water_id:
        water_period = WaterReadouts.objects.get(pk=water_id)
        water_date = water_period.readout_date
        water_usage = WaterReadouts.objects.filter(apartment=apt, readout_date__lte=water_date).order_by('-readout_date')
    else:
        water_usage = WaterReadouts.objects.filter(apartment=apt).order_by('-readout_date')
    cold, hot, cold_curr, hot_curr = calculate_water_usage(water_usage)

    fees = Fees.objects.values().filter(period__lte=water_usage[0].readout_date).latest('period')

    readout_date = water_usage[0].readout_date
    tenants = apt.occupancy_set.filter(start_date__lte=readout_date).latest('start_date').occupants
    parking = apt.parkingcard_set.filter(start_date__lte=readout_date).order_by('start_date')
    big_family_card = apt.bigfamilycard_set.filter(start_date__lte=readout_date).order_by('start_date')
    ch_surcharge = apt.centralheatingsurcharge_set.filter(start_date__lte=readout_date, end_date__gte=readout_date)
    ch_payment = 0
    if ch_surcharge:
        if readout_date == ch_surcharge[0].end_date:
            ch_payment = ch_surcharge[0].last_payment
        else:
            ch_payment = ch_surcharge[0].amount_per_month

    parking_cards = parking[0].number_of_cards if parking else 0
    obj = {'readout_date': readout_date,
           'hot_water_cost': hot * fees['hot_water'],
           'cold_water_cost': cold * fees['cold_water'],
           'maintenance_cost': apt.area * fees['maintenance_fee'],
           'repair_fund_cost': apt.area * fees['repair_fund'],
           'central_heating_cost': apt.area * fees['central_heating'],
           'garbage_cost': tenants * fees['garbage'],
           'tenants': tenants,
           'hot_used': hot,
           'hot_water': hot_curr,
           'cold_used': cold,
           'cold_water': cold_curr,
           'hot_water_fee': fees['hot_water'],
           'cold_water_fee': fees['cold_water'],
           'maintenance_fee': fees['maintenance_fee'],
           'repair_fund_fee': fees['repair_fund'],
           'central_heating_fee': fees['central_heating'],
           'garbage_fee': fees['garbage'],
           'parking_fee': parking_cards * fees['parking_fee'],
           'central_heating_surcharge': ch_payment,
           }
    if big_family_card:
        obj['garbage_cost'] -= big_family_card[0].amount
    obj['total'] = obj['hot_water_cost'] + obj['cold_water_cost'] + obj['maintenance_cost'] + obj['repair_fund_cost'] + obj['central_heating_cost'] + obj['garbage_cost'] + obj['parking_fee'] + obj['central_heating_surcharge']

    return obj


def import_water_readouts(file):
    wb = load_workbook(filename=file)
    sheet = wb.active
    dates = [date for date in sheet.iter_rows(max_row=1, values_only=True)]
    dates = [date for date in dates[0] if date]
    water_readouts = [water for water in sheet.iter_rows(min_row=2, values_only=True)]

    for water_readout in water_readouts:
        apt = water_readout[0]
        if apt:
            cold_water = water_readout[1::2]
            hot_water = water_readout[2::2]
            water = zip(cold_water, hot_water)
            for value, date in enumerate(dates):
                if date:
                    obj = WaterReadouts.objects.create(apartment=Apartment.objects.get(number=apt),
                                                       readout_date=date,
                                                       cold_water_readout=cold_water[value],
                                                       hot_water_readout=hot_water[value],
                                                       new_hot_water_meter=False,
                                                       new_cold_water_meter=False)
